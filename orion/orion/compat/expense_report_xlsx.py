"""Workbook builder for the Laporan Pengeluaran Proyek .xlsx export.

Pure openpyxl — no frappe imports — so the layout can be exercised and
diffed against the reference workbook outside a bench. The layout mirrors
the hand-styled LPP-R-KN02-LCTR-Feedback.xlsx: Ratunda letterhead (logo +
address), purple 7030A0 title bar and table header, lavender (accent4
tint 0.8) info block with bold-label / plain-value rich text, fully
bordered line table, TOTAL row, and CREATED/AUTHORIZED signature boxes.

`build_workbook(data, meta, logo_path)` takes the _report_read payload
plus display metadata resolved by the caller:

  meta = {
      "brand":         "RATUNDA RENOVASI" | "POIESIS STUDIO" | ...,
      "projectName":   str,
      "clientName":    str,
      "location":      str,
      "projectStatus": str,  # AKTIF / SELESAI / DITUNDA / BATAL
  }
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Border, Color, Font, PatternFill, Protection, Side
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

PURPLE = "FF7030A0"
WHITE = Color(theme=0)
BLACK = Color(theme=1)
LAVENDER = Color(theme=7, tint=0.8)  # accent4 (8064A2) lightened 80%

ADDRESS = "Calamus C7/35. Citra Garden Bintaro. Tangerang Selatan"

SHEETS = (
	("PROYEK", "REKAP PROYEK", "LAPORAN KEUANGAN PROYEK - REKAP"),
	("BON_MATERIAL", "REKAP BON MATERIAL", "LAPORAN KEUANGAN PROYEK - BON MATERIAL"),
	("UPAH", "REKAP UPAH", "LAPORAN KEUANGAN PROYEK - UPAH PEKERJA"),
)
HEADERS = [
	"DATE", "INVOICE NO", "PURCHASE NO", "ITEM", "ITEM DESCRIPTION",
	"VENDOR", "QTY", "DEBIT", "CREDIT", "AMOUNT",
]
# content lives in B..K; A is a 2-wide gutter
COLS = "BCDEFGHIJK"
WIDTHS = {"A": 2, "B": 12, "C": 15.88, "D": 17.12, "E": 22, "F": 34, "G": 20, "H": 6, "I": 15, "J": 15, "K": 16}

MONTHS_ID = [
	"Januari", "Februari", "Maret", "April", "Mei", "Juni",
	"Juli", "Agustus", "September", "Oktober", "November", "Desember",
]

THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# "Rp 10.000" — the thousands separator renders per Excel locale (dot on id-ID)
RP_FMT = '"Rp "#,##0'

# Template mode: sheet-protected fill-in form. The password is tamper
# discouragement for the field team, not security.
TEMPLATE_ROWS = 150
TEMPLATE_PASSWORD = "RATUNDA2026"
TEMPLATE_MARKER = "orion-lpp-template-v1"
UNLOCKED = Protection(locked=False)


def _iso_to_date(iso: str | None):
	try:
		return datetime.strptime((iso or "")[:10], "%Y-%m-%d").date()
	except ValueError:
		return None


def _date_id(iso: str | None) -> str:
	"""'2026-08-05T…' -> '5 Agustus 2026' (empty on missing/garbled input)."""
	try:
		y, m, d = (iso or "")[:10].split("-")
		if not 1 <= int(m) <= 12:
			return ""
		return "%d %s %s" % (int(d), MONTHS_ID[int(m) - 1], y)
	except (ValueError, IndexError):
		return ""


def _label_value(ws, coord: str, label: str, value: str):
	"""Rich text 'LABEL: value' — label run inherits the bold cell font,
	value run is written plain (Verdana 11, black)."""
	c = ws[coord]
	c.font = Font(name="Verdana", size=11, bold=True, color=BLACK)
	if value:
		c.value = CellRichText(
			"%s: " % label,
			TextBlock(InlineFont(rFont="Verdana", sz=11, color=BLACK), value),
		)
	else:
		c.value = "%s: " % label


def _sig_row(ws, row: int, label: str, value: str):
	"""One line of the signature block: bordered G..K strip, label in G
	(white-filled) divided from the bold value in H."""
	white = PatternFill("solid", fgColor=WHITE)
	ws.cell(row=row, column=7, value=label).font = Font(name="Verdana", size=10, color=BLACK)
	ws.cell(row=row, column=7).fill = white
	ws.cell(row=row, column=7).border = Border(left=THIN, top=THIN, bottom=THIN)
	ws.cell(row=row, column=8, value=value or None).font = Font(name="Verdana", size=10, bold=True, color=BLACK)
	ws.cell(row=row, column=8).border = Border(left=THIN, top=THIN, bottom=THIN)
	for col in (9, 10):
		ws.cell(row=row, column=col).border = Border(top=THIN, bottom=THIN)
	ws.cell(row=row, column=11).border = Border(right=THIN, top=THIN, bottom=THIN)


def _add_logo(ws, logo_path: str):
	# sized/offset as in the reference workbook: ~163x45 px, nudged inside B1
	from openpyxl.drawing.image import Image
	from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
	from openpyxl.drawing.xdr import XDRPositiveSize2D

	img = Image(logo_path)
	img.anchor = OneCellAnchor(
		_from=AnchorMarker(col=1, colOff=47625, row=0, rowOff=57150),
		ext=XDRPositiveSize2D(cx=1553845, cy=426085),
	)
	ws.add_image(img)


def _build_sheet(ws, data: dict, key: str, title: str, meta: dict, logo_path: str | None, template: bool = False):
	white = PatternFill("solid", fgColor=WHITE)
	purple = PatternFill("solid", fgColor=PURPLE)
	lavender = PatternFill("solid", fgColor=LAVENDER)

	for col, width in WIDTHS.items():
		ws.column_dimensions[col].width = width
	ws.row_dimensions[1].height = 37
	ws.row_dimensions[4].height = 18

	# letterhead: logo row + address over white, ruled off below the address
	for row in (1, 2, 3):
		for col in COLS:
			ws["%s%d" % (col, row)].fill = white
	if logo_path:
		_add_logo(ws, logo_path)
	ws["B2"] = ADDRESS
	ws["B2"].font = Font(name="Arial", size=10, color=BLACK)
	for col in COLS:
		ws["%s2" % col].border = Border(bottom=THIN)

	# purple title bar
	ws["B4"] = title
	ws["B4"].font = Font(name="Verdana", size=14, bold=True, color=WHITE)
	for col in COLS:
		ws["%s4" % col].fill = purple

	# lavender info block
	for row in (5, 6, 7, 8):
		for col in COLS:
			ws["%s%d" % (col, row)].fill = lavender
	ws["B5"] = meta.get("brand") or ""
	ws["B5"].font = Font(name="Verdana", size=11, bold=True, color=BLACK)
	_label_value(ws, "B6", "PROYEK", meta.get("projectName") or "")
	_label_value(ws, "E6", "CLIENT", meta.get("clientName") or "")
	_label_value(ws, "B7", "KODE", data.get("code") or "")
	_label_value(ws, "E7", "LOKASI", meta.get("location") or "")
	_label_value(ws, "B8", "STATUS PROYEK", meta.get("projectStatus") or "")

	# line table: purple header row 10, bordered data rows, TOTAL right below
	header_font = Font(name="Verdana", size=10, bold=True, color=WHITE)
	body_font = Font(name="Verdana", size=10, color=BLACK)
	for col, text in zip(COLS, HEADERS):
		c = ws["%s10" % col]
		c.value = text
		c.font = header_font
		c.fill = purple
		c.border = BOX

	sec = data["sections"][key]
	row = 10
	lines = sec["lines"]
	count = max(len(lines), TEMPLATE_ROWS) if template else len(lines)
	for i in range(count):
		row += 1
		ln = lines[i] if i < len(lines) else None
		if ln is None:
			values = [None] * 10
		elif template:
			# real date + only B..J values; K becomes a running-saldo formula
			values = [
				_iso_to_date(ln["date"]),
				ln["invoiceNo"], ln["purchaseNo"], ln["item"], ln["itemDescription"], ln["vendor"],
				ln["qty"], float(ln["debit"]), float(ln["credit"]),
			]
		else:
			values = [
				(ln["date"] or "")[:10],
				ln["invoiceNo"], ln["purchaseNo"], ln["item"], ln["itemDescription"], ln["vendor"],
				ln["qty"], float(ln["debit"]), float(ln["credit"]), float(ln["amount"]),
			]
		for col, value in zip(COLS, values):
			ws["%s%d" % (col, row)].value = value
		for col in COLS:
			c = ws["%s%d" % (col, row)]
			c.font = body_font
			c.border = BOX
			if col in "IJK":
				c.number_format = RP_FMT
			if template:
				if col == "B":
					c.number_format = "dd/mm/yyyy"
				elif col in "CDEFG":
					c.number_format = "@"  # force text: keeps refs like 01-02 out of date coercion
				if col != "K":
					c.protection = UNLOCKED
		if template:
			# live running saldo; N() zeroes the header-row anchor
			ws["K%d" % row] = '=IF(AND(I{r}="",J{r}=""),"",N(K{p})+N(I{r})-N(J{r}))'.format(r=row, p=row - 1)

	row += 1
	total_font = Font(name="Verdana", size=10, bold=True, color=BLACK)
	if template:
		totals = {
			"G": "TOTAL", "H": None,
			"I": "=SUM(I11:I%d)" % (row - 1),
			"J": "=SUM(J11:J%d)" % (row - 1),
			"K": "=I%d-J%d" % (row, row),
		}
	else:
		totals = {"G": "TOTAL", "H": None, "I": float(sec["totalDebit"]), "J": float(sec["totalCredit"]), "K": float(sec["balance"])}
	for col, value in totals.items():
		c = ws["%s%d" % (col, row)]
		c.value = value
		c.font = total_font
		c.border = BOX
		if col in "IJK":
			c.number_format = RP_FMT
	ws["G%d" % row].fill = white

	# signature boxes
	authorized = data.get("status") == "AUTHORIZED"
	_sig_row(ws, row + 2, "CREATED BY", data.get("preparedBy") or "")
	_sig_row(ws, row + 3, "CREATED ON", _date_id(data.get("createdAt")))
	_sig_row(ws, row + 5, "AUTHORIZED BY", data.get("authorizedBy") or "")
	_sig_row(ws, row + 6, "AUTHORIZED ON", _date_id(data.get("updatedAt")) if authorized else "")

	if template:
		last = 10 + count
		dv_date = DataValidation(
			type="date", operator="between", formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)",
			allow_blank=True, showErrorMessage=True,
			errorTitle="Tanggal", error="Isi tanggal yang valid (dd/mm/yyyy)",
		)
		dv_num = DataValidation(
			type="decimal", operator="greaterThanOrEqual", formula1="0",
			allow_blank=True, showErrorMessage=True,
			errorTitle="Angka", error="Isi angka saja, tanpa Rp dan tanpa titik ribuan",
		)
		ws.add_data_validation(dv_date)
		ws.add_data_validation(dv_num)
		dv_date.add("B11:B%d" % last)
		dv_num.add("H11:J%d" % last)  # QTY + DEBIT + CREDIT
		# lock everything except the unlocked data cells; row insertion stays allowed
		ws.protection = SheetProtection(
			sheet=True, insertRows=False, selectLockedCells=False, selectUnlockedCells=False,
			formatCells=True, formatColumns=True, formatRows=True,
			deleteRows=True, deleteColumns=True, insertColumns=True, sort=True, autoFilter=True,
		)
		ws.protection.set_password(TEMPLATE_PASSWORD)


def build_workbook(data: dict, meta: dict, logo_path: str | None, template: bool = False) -> Workbook:
	wb = Workbook()
	wb.remove(wb.active)
	for key, sheet, title in SHEETS:
		_build_sheet(wb.create_sheet(sheet), data, key, title, meta, logo_path, template)
	if template:
		wb.security = WorkbookProtection(lockStructure=True)
		wb.security.workbook_password = TEMPLATE_PASSWORD
		wb.properties.keywords = TEMPLATE_MARKER
	return wb
