"""Project expense report ("Laporan Pengeluaran Proyek") compat handlers.

A per-project cash-management ledger the project architect fills in — the
digital version of the LaporanKeuanganProyek Excel (three tabs: REKAP PROYEK,
REKAP BON MATERIAL, REKAP UPAH). Each line can be TAGGED to an existing Cash
Advance (Orion Cash Advance, CAM/CAO/CAU) and, for material rows, a Nota
(Orion Field Receipt).

This is a NON-GL tracking ledger: it posts NO Journal Entry. GL truth stays
with the linked Cash Advances / client Payment Entries; this report is a cash
working view + advance-vs-receipts reconciliation. Its totals must never be
summed into project-financial / GL figures (they overlap the linked docs).

Routes served through the JSON gateway (orion.compat.handle):

  GET   /api/supply-chain/expense-reports?projectId=  -> list
  GET   /api/supply-chain/expense-reports/:id         -> detail (sections with
        server-computed running balance + per-tab totals + margin + CA/Nota stubs)
  POST  /api/supply-chain/expense-reports             -> get-or-create per project
  PATCH /api/supply-chain/expense-reports/:id         -> replace header + lines (DRAFT only)
  POST  /api/supply-chain/expense-reports/:id/authorize -> DRAFT -> AUTHORIZED

Binary export cannot ride the JSON gateway; the .xlsx download is a dedicated
whitelisted method hit directly by feynman (GET, session cookie):

  GET /api/method/orion.compat.expense_report.export_xlsx?reportId=<id>

Money serializes as 2dp JSON strings (matching accounting.py / supply_chain.py);
ids resolve orion_legacy_id-first and emit orion_legacy_id when present.
"""

from decimal import Decimal

import frappe

from orion.compat.accounting import (
	_iso,
	_iso_date,
	_legacy_or_name,
	_money,
	_parse_date,
	_resolve_name,
	_s2,
	_wdec,
)
from orion.compat.handle import route, split_path
from orion.compat.supply_chain import _ca_stubs, _project_stubs

PREFIX = "/api/supply-chain/expense-reports"
DOCTYPE = "Orion Project Expense Report"
LINE_DOCTYPE = "Orion Project Expense Line"
SECTIONS = ("PROYEK", "BON_MATERIAL", "UPAH")


@route(PREFIX)
def expense_reports(path: str, verb: str, payload: dict):
	bare, query = split_path(path)
	rest = bare[len(PREFIX):].strip("/")
	parts = [p for p in rest.split("/") if p]
	if verb == "GET" and not parts:
		return _report_list(query)
	if verb == "POST" and not parts:
		return _report_create(payload)
	if len(parts) == 1:
		if verb == "GET":
			return _report_read(_report_name(parts[0]))
		if verb == "PATCH":
			return _report_update(parts[0], payload)
		if verb == "DELETE":
			return _report_delete(parts[0])
	if len(parts) == 2 and verb == "POST" and parts[1] == "authorize":
		return _report_transition(parts[0], "DRAFT", "AUTHORIZED", "Only DRAFT reports can be authorized")
	frappe.throw("No compat handler for %s %s" % (verb, bare), exc=frappe.DoesNotExistError)


def _report_name(ident: str) -> str:
	name = _resolve_name(DOCTYPE, ident)
	if not name:
		frappe.throw("Not found", exc=frappe.DoesNotExistError)
	return name


def _fr_stubs(names: set) -> dict:
	if not names:
		return {}
	return {
		r.name: {
			"id": r.orion_legacy_id or r.name,
			"code": r.name,
			"storeName": r.store_name,
			"totalAmount": _s2(r.total_amount),
			"status": r.status,
		}
		for r in frappe.get_all(
			"Orion Field Receipt",
			filters={"name": ("in", list(names))},
			fields=["name", "store_name", "total_amount", "status", "orion_legacy_id"],
		)
	}


def _report_read(name: str) -> dict:
	d = frappe.get_doc(DOCTYPE, name)
	rows = sorted(d.lines, key=lambda r: (r.section or "PROYEK", r.sort_order or 0, str(r.line_date or "")))
	ca_stub = _ca_stubs({r.cash_advance for r in d.lines if r.cash_advance})
	fr_stub = _fr_stubs({r.field_receipt for r in d.lines if r.field_receipt})

	sections = {s: {"lines": [], "debit": Decimal("0"), "credit": Decimal("0")} for s in SECTIONS}
	running = {s: Decimal("0") for s in SECTIONS}
	for r in rows:
		s = r.section if r.section in SECTIONS else "PROYEK"
		deb, cred = _wdec(r.debit), _wdec(r.credit)
		running[s] += cred - deb  # CREDIT (inflow) adds, DEBIT (outflow) draws down
		sections[s]["debit"] += deb
		sections[s]["credit"] += cred
		sections[s]["lines"].append(
			{
				"id": r.name,
				"section": s,
				"date": _iso_date(r.line_date),
				"invoiceNo": r.invoice_no,
				"purchaseNo": r.purchase_no,
				"item": r.item,
				"itemDescription": r.item_description,
				"vendor": r.vendor,
				"qty": r.qty,
				"debit": _s2(r.debit),
				"credit": _s2(r.credit),
				"amount": _s2(running[s]),  # running balance, per section
				"cashAdvanceId": _legacy_or_name("Orion Cash Advance", r.cash_advance) if r.cash_advance else None,
				"fieldReceiptId": _legacy_or_name("Orion Field Receipt", r.field_receipt) if r.field_receipt else None,
				"cashAdvance": ca_stub.get(r.cash_advance),
				"fieldReceipt": fr_stub.get(r.field_receipt),
				"notes": r.notes,
				"sortOrder": r.sort_order or 0,
			}
		)

	proyek = sections["PROYEK"]
	net = proyek["credit"] - proyek["debit"]
	margin = float(net / proyek["credit"] * 100) if proyek["credit"] else 0.0
	return {
		"id": d.orion_legacy_id or d.name,
		"code": d.code,
		"projectId": _legacy_or_name("Project", d.project) if d.project else None,
		"project": _project_stubs({d.project}).get(d.project) if d.project else None,
		"status": d.status,
		"title": d.title,
		"businessLine": d.business_line,
		"preparedBy": d.prepared_by,
		"authorizedBy": d.authorized_by,
		"notes": d.notes,
		"sections": {
			s: {
				"lines": v["lines"],
				"totalDebit": _s2(v["debit"]),
				"totalCredit": _s2(v["credit"]),
				"balance": _s2(v["credit"] - v["debit"]),
			}
			for s, v in sections.items()
		},
		"marginPct": round(margin, 2),
		"createdAt": _iso(d.creation),
		"updatedAt": _iso(d.modified),
	}


def _report_list(query: dict) -> list:
	filters = {}
	if query.get("projectId"):
		pn = _resolve_name("Project", query["projectId"])
		filters["project"] = pn or "__no_such_project__"
	if query.get("status"):
		filters["status"] = query["status"]
	names = frappe.get_all(DOCTYPE, filters=filters, order_by="modified desc", pluck="name")
	return [_report_read(n) for n in names]


def _report_code(project_name: str) -> str:
	p = frappe.db.get_value(
		"Project", project_name, ["orion_business_line", "kn_sequence", "lctr"], as_dict=True
	) or frappe._dict()
	bl = "R" if (p.orion_business_line or "RATUNDA_RENOVASI") == "RATUNDA_RENOVASI" else "P"
	return "LPP-%s-KN%02d-%s" % (bl, int(p.kn_sequence or 0), p.lctr or "XXXX")


def _line_row(ln: dict, idx: int) -> dict:
	sec = ln.get("section") if ln.get("section") in SECTIONS else "PROYEK"
	return {
		"section": sec,
		"line_date": _parse_date(ln.get("date")),
		"invoice_no": ln.get("invoiceNo"),
		"purchase_no": ln.get("purchaseNo"),
		"item": ln.get("item"),
		"item_description": ln.get("itemDescription"),
		"vendor": ln.get("vendor"),
		"qty": float(ln["qty"]) if ln.get("qty") not in (None, "") else None,
		"debit": _money(ln.get("debit")),
		"credit": _money(ln.get("credit")),
		"cash_advance": _resolve_name("Orion Cash Advance", ln.get("cashAdvanceId")),
		"field_receipt": _resolve_name("Orion Field Receipt", ln.get("fieldReceiptId")),
		"notes": ln.get("notes"),
		"sort_order": idx,
	}


def _report_create(payload: dict) -> dict:
	project = _resolve_name("Project", payload.get("projectId"))
	if not project:
		frappe.throw("Project not found", exc=frappe.DoesNotExistError)
	existing = frappe.db.get_value(DOCTYPE, {"project": project})
	if existing:
		return _report_read(existing)  # idempotent: one report per project

	bl = frappe.db.get_value("Project", project, "orion_business_line") or "RATUNDA_RENOVASI"
	d = frappe.new_doc(DOCTYPE)
	d.code = _report_code(project)
	d.project = project
	d.business_line = bl
	d.status = "DRAFT"
	d.title = payload.get("title")
	d.prepared_by = payload.get("preparedBy")
	d.authorized_by = payload.get("authorizedBy")
	d.notes = payload.get("notes")
	for idx, ln in enumerate(payload.get("lines") or []):
		d.append("lines", _line_row(ln, idx))
	d.flags.ignore_permissions = True
	d.insert()
	return _report_read(d.name)


def _report_update(ident: str, payload: dict) -> dict:
	d = frappe.get_doc(DOCTYPE, _report_name(ident))
	if d.status != "DRAFT":
		frappe.throw("Only DRAFT reports can be edited")
	for src, dst in [
		("title", "title"),
		("preparedBy", "prepared_by"),
		("authorizedBy", "authorized_by"),
		("notes", "notes"),
	]:
		if src in payload:
			setattr(d, dst, payload[src])
	if payload.get("lines") is not None:
		d.set("lines", [])
		for idx, ln in enumerate(payload["lines"]):
			d.append("lines", _line_row(ln, idx))
	d.flags.ignore_permissions = True
	d.save()
	return _report_read(d.name)


def _report_delete(ident: str) -> dict:
	d = frappe.get_doc(DOCTYPE, _report_name(ident))
	if d.status != "DRAFT":
		frappe.throw("Only DRAFT reports can be deleted")
	frappe.delete_doc(DOCTYPE, d.name, force=1, ignore_permissions=True)
	return {"ok": True}


def _report_transition(ident: str, need: str, to: str, message: str) -> dict:
	d = frappe.get_doc(DOCTYPE, _report_name(ident))
	if d.status != need:
		frappe.throw(message)
	d.status = to
	d.flags.ignore_permissions = True
	d.save()
	return _report_read(d.name)


@frappe.whitelist(methods=["GET"])
def export_xlsx(reportId: str):
	"""Direct binary download (the JSON gateway can't stream files). Builds the
	three-sheet LaporanKeuanganProyek workbook from _report_read (server-computed
	running balances) and returns it via frappe.response, like build_xlsx_response."""
	if frappe.session.user in ("Guest", "", None):
		raise frappe.AuthenticationError
	name = _resolve_name(DOCTYPE, reportId)
	if not name:
		frappe.throw("Not found", exc=frappe.DoesNotExistError)
	data = _report_read(name)

	from io import BytesIO

	from openpyxl import Workbook
	from openpyxl.styles import Font

	brand = {
		"RATUNDA_RENOVASI": "RATUNDA",
		"POIESIS_STUDIO": "POIESIS",
	}.get(data.get("businessLine"), "RATUNDA / POIESIS")
	proj = data.get("project") or {}
	headers = [
		"DATE", "INVOICE NO", "PURCHASE NO", "ITEM", "ITEM DESCRIPTION",
		"VENDOR", "QTY", "DEBIT", "CREDIT", "AMOUNT",
	]
	widths = [12, 12, 12, 22, 34, 20, 6, 15, 15, 16]

	wb = Workbook()
	wb.remove(wb.active)
	for key, sheet in (("PROYEK", "REKAP PROYEK"), ("BON_MATERIAL", "REKAP BON MATERIAL"), ("UPAH", "REKAP UPAH")):
		ws = wb.create_sheet(sheet)
		ws["A1"] = sheet
		ws["A1"].font = Font(bold=True, size=14)
		ws["A2"] = brand
		ws["A3"] = "PROYEK: %s" % (proj.get("name") or "")
		ws["A4"] = "KODE: %s" % data["code"]
		ws.append([])
		ws.append(headers)
		for c in ws[6]:
			c.font = Font(bold=True)
		sec = data["sections"][key]
		for ln in sec["lines"]:
			ws.append([
				(ln["date"] or "")[:10],
				ln["invoiceNo"], ln["purchaseNo"], ln["item"], ln["itemDescription"], ln["vendor"],
				ln["qty"], float(ln["debit"]), float(ln["credit"]), float(ln["amount"]),
			])
		ws.append([])
		total_row = ["", "", "", "", "", "TOTAL", "", float(sec["totalDebit"]), float(sec["totalCredit"]), float(sec["balance"])]
		ws.append(total_row)
		for c in ws[ws.max_row]:
			c.font = Font(bold=True)
		for i, w in enumerate(widths):
			ws.column_dimensions[chr(ord("A") + i)].width = w

	buf = BytesIO()
	wb.save(buf)
	frappe.response["filename"] = "%s.xlsx" % data["code"]
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"
