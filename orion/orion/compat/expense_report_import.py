"""Parser for uploaded LaporanKeuanganProyek workbooks.

Pure openpyxl — no frappe imports — so it can be exercised against the real
historical files outside a bench. Understands both the current Orion export
layout (header DATE…AMOUNT in one row, one DEBIT/CREDIT pair) and the legacy
hand-made LaporanKeuanganProyek files (header around row 12, a dummy money
pair in D/E plus the real pair in L/M, running AMOUNT in N, TOTAL footer,
signature block).

Sign convention: in Orion, Saldo = DEBIT − CREDIT (DEBIT adds). Legacy files
label the adding column inconsistently (CREDIT on REKAP PROYEK, DEBIT on BON
MATERIAL/UPAH), so the parser orients the money pair by *balance effect* —
which column moves the running AMOUNT up — and only falls back to labels when
no AMOUNT column exists. A swapped sheet is reported in `sheets[].swapped`
and as a warning so the preview can say so.

parse_workbook(content) -> {
    "sections": {"PROYEK": [line], "BON_MATERIAL": [...], "UPAH": [...]},
    "sheets":   [{"name", "section", "rows", "headerRow", "swapped"}],
    "warnings": [str],
}
line = {section, date (ISO|None), invoiceNo, purchaseNo, item,
        itemDescription, vendor, qty, debit, credit}
"""

import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

SECTIONS = ("PROYEK", "BON_MATERIAL", "UPAH")

MONEY_FIELDS = ("debit", "credit", "amount")
TEXT_FIELDS = ("invoiceNo", "purchaseNo", "item", "itemDescription", "vendor")

STOP_MARKERS = {"TOTAL", "SISA SALDO", "GRAND TOTAL", "TOTAL KESELURUHAN"}

DATE_PATTERNS = (
	("%Y-%m-%d", re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")),
	("%d/%m/%Y", re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")),
	("%d-%m-%Y", re.compile(r"^\d{1,2}-\d{1,2}-\d{4}$")),
	("%d.%m.%Y", re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$")),
	("%d/%m/%y", re.compile(r"^\d{1,2}/\d{1,2}/\d{2}$")),
)


def _norm(v) -> str:
	return re.sub(r"[\s.:_]+", " ", str(v)).strip().upper() if v is not None else ""


def _header_field(v):
	t = _norm(v)
	if not t:
		return None
	if t in ("DATE", "TGL", "TANGGAL"):
		return "date"
	if t.startswith("INVOICE") or t in ("NO INVOICE", "INV NO"):
		return "invoiceNo"
	if t.startswith("PURCHASE") or t in ("PO", "PO NO", "NO PO"):
		return "purchaseNo"
	if "DESCRIPTION" in t or "DESKRIPSI" in t or "KETERANGAN" in t or "URAIAN" in t:
		return "itemDescription"
	if t.startswith("ITEM") or t == "BARANG":
		return "item"
	if "VENDOR" in t or "PEMBELI" in t or "PENERIMA" in t or t in ("TOKO", "SUPPLIER", "ACTION BY"):
		return "vendor"
	if t in ("QTY", "QUANTITY", "JML"):
		return "qty"
	if t in ("DEBIT", "DEBET"):
		return "debit"
	if t in ("CREDIT", "KREDIT"):
		return "credit"
	if t in ("AMOUNT", "SALDO", "BALANCE", "SISA SALDO"):
		return "amount"
	return None


def _sheet_section(title: str) -> str:
	t = _norm(title)
	if "BON" in t or "MATERIAL" in t:
		return "BON_MATERIAL"
	if "UPAH" in t or "TUKANG" in t or "WAGE" in t:
		return "UPAH"
	return "PROYEK"


def _parse_money(v):
	if v is None or isinstance(v, bool):
		return None
	if isinstance(v, (int, float)):
		return float(v)
	s = str(v).strip()
	if not s or s in ("-", "—"):
		return None
	neg = s.startswith("(") and s.endswith(")")
	s = re.sub(r"(?i)^rp\.?", "", s.strip("() ")).replace("\xa0", "").replace(" ", "")
	if not s:
		return None
	if "." in s and "," in s:  # last separator is the decimal one
		if s.rfind(",") > s.rfind("."):
			s = s.replace(".", "").replace(",", ".")
		else:
			s = s.replace(",", "")
	elif re.fullmatch(r"\d{1,3}(\.\d{3})+", s):  # 1.234.567 — Indonesian thousands
		s = s.replace(".", "")
	elif "," in s:
		s = s.replace(",", ".") if re.fullmatch(r"\d+,\d{1,2}", s) else s.replace(",", "")
	try:
		n = float(s)
	except ValueError:
		return None
	return -n if neg else n


def _parse_row_date(v):
	if isinstance(v, datetime):
		return v.date().isoformat()
	if isinstance(v, date):
		return v.isoformat()
	if isinstance(v, str):
		s = v.strip()[:10]
		for fmt, pat in DATE_PATTERNS:
			if pat.match(s):
				try:
					return datetime.strptime(s, fmt).date().isoformat()
				except ValueError:
					return None
	return None


def _find_header(ws):
	"""(header_row, {field: col | [cols] for money fields}) — a header row has
	DATE plus at least one money column and 4+ recognized fields."""
	for row in ws.iter_rows(min_row=1, max_row=min(40, ws.max_row)):
		fields = {}
		for c in row:
			f = _header_field(c.value)
			if not f:
				continue
			if f in MONEY_FIELDS:
				fields.setdefault(f, []).append(c.column)
			else:
				fields.setdefault(f, c.column)
		if "date" in fields and ("debit" in fields or "credit" in fields) and len(fields) >= 4:
			return row[0].row, fields
	return None, None


def _pick_money_col(cols, rows, key):
	"""Of duplicate DEBIT/CREDIT/AMOUNT columns keep the one actually carrying
	numbers (legacy files have a dummy pair in D/E); ties go rightmost."""
	if not cols:
		return None
	best, best_n = None, -1
	for col in cols:
		n = sum(1 for r in rows if r[key].get(col) is not None)
		if n >= best_n:
			best, best_n = col, n
	return best


def _orient(rows, debit_col, credit_col, amount_col):
	"""True = file DEBIT column is the one that grows the balance (Orion
	debit); decided by votes from the running AMOUNT column."""
	votes = 0  # >0 direct, <0 swapped
	prev = 0.0
	for r in rows:
		amount = r["money"].get(amount_col) if amount_col else None
		if amount is None:
			continue
		deb = (r["money"].get(debit_col) if debit_col else None) or 0.0  # 0/None = not filled
		cred = (r["money"].get(credit_col) if credit_col else None) or 0.0
		val, prev_amount = deb or cred, prev
		prev = amount
		if bool(deb) == bool(cred):  # both/neither filled: no signal
			continue
		# vote only on exact running-balance steps — some legacy sheets put a
		# per-row copy of the value in AMOUNT instead of a running total
		if abs(amount - prev_amount - val) < 0.01:  # this side grew the balance
			votes += 1 if deb else -1
		elif abs(amount - prev_amount + val) < 0.01:  # this side shrank it
			votes += -1 if deb else 1
	return votes >= 0, votes != 0


def _parse_sheet(ws):
	header_row, fields = _find_header(ws)
	if not header_row:
		return None
	debit_cols = fields.get("debit", [])
	credit_cols = fields.get("credit", [])
	amount_cols = fields.get("amount", [])

	raw_rows = []
	empty_streak = 0
	for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
		cells = {c.column: c.value for c in row if c.value is not None}
		if any(_norm(v) in STOP_MARKERS for v in cells.values()):
			break
		if not cells:
			empty_streak += 1
			if empty_streak >= 20:
				break
			continue
		empty_streak = 0
		raw_rows.append(
			{
				"date": _parse_row_date(cells.get(fields["date"])),
				"text": {f: cells.get(fields[f]) for f in TEXT_FIELDS if f in fields},
				"qty": _parse_money(cells.get(fields["qty"])) if "qty" in fields else None,
				"money": {col: _parse_money(cells.get(col)) for col in set(debit_cols + credit_cols + amount_cols)},
			}
		)

	debit_col = _pick_money_col(debit_cols, raw_rows, "money")
	credit_col = _pick_money_col(credit_cols, raw_rows, "money")
	amount_col = _pick_money_col(amount_cols, raw_rows, "money")

	rows = [
		r for r in raw_rows
		if r["date"] is not None
		or (debit_col and r["money"].get(debit_col) is not None)
		or (credit_col and r["money"].get(credit_col) is not None)
	]
	direct, decided = _orient(rows, debit_col, credit_col, amount_col)
	if not decided:
		# no usable AMOUNT column: legacy labels the growing column CREDIT on
		# REKAP PROYEK and DEBIT on BON MATERIAL/UPAH
		direct = _sheet_section(ws.title) != "PROYEK"
	if not direct:
		debit_col, credit_col = credit_col, debit_col

	lines = []
	for r in rows:
		text = {f: (str(v).strip() or None) if v is not None else None for f, v in r["text"].items()}
		lines.append(
			{
				"date": r["date"],
				"invoiceNo": text.get("invoiceNo"),
				"purchaseNo": text.get("purchaseNo"),
				"item": text.get("item"),
				"itemDescription": text.get("itemDescription"),
				"vendor": text.get("vendor"),
				"qty": r["qty"],
				"debit": r["money"].get(debit_col) if debit_col else None,
				"credit": r["money"].get(credit_col) if credit_col else None,
			}
		)
	return {"headerRow": header_row, "swapped": not direct, "lines": lines}


def parse_workbook(content: bytes) -> dict:
	wb = load_workbook(BytesIO(content), data_only=True, read_only=False)
	sections = {s: [] for s in SECTIONS}
	sheets, warnings = [], []
	taken = {}
	for ws in wb.worksheets:
		if ws.sheet_state != "visible":
			continue
		parsed = _parse_sheet(ws)
		if not parsed:
			continue
		section = _sheet_section(ws.title)
		if section in taken:
			warnings.append("Sheet '%s' dilewati — tab %s sudah terisi dari sheet '%s'" % (ws.title, section, taken[section]))
			continue
		taken[section] = ws.title
		for ln in parsed["lines"]:
			ln["section"] = section
		sections[section] = parsed["lines"]
		if parsed["swapped"]:
			warnings.append(
				"Sheet '%s': kolom CREDIT file dibaca sebagai DEBIT Orion (kas masuk) agar saldo cocok" % ws.title
			)
		sheets.append(
			{
				"name": ws.title,
				"section": section,
				"rows": len(parsed["lines"]),
				"headerRow": parsed["headerRow"],
				"swapped": parsed["swapped"],
			}
		)
	return {"sections": sections, "sheets": sheets, "warnings": warnings}
